from flask import Flask, render_template, request, redirect, session, url_for, send_file, flash
import pymysql
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import math
from functools import wraps
import os
import re

app = Flask(__name__)
app.secret_key = 'hoteloasis2024'

def conectar():
    # Para Railway (usando MYSQL_URL)
    mysql_url = os.environ.get('MYSQL_URL')
    print(f"DEBUG: MYSQL_URL = {mysql_url}") 
    if mysql_url:
        # Parsear la URL que Railway genera
        patron = r'mysql://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)'
        match = re.match(patron, mysql_url)
        if match:
            usuario = match.group(1)
            password = match.group(2)
            host = match.group(3)
            puerto = int(match.group(4))
            bd = match.group(5)
            
            return pymysql.connect(
                host=host,
                user=usuario,
                password=password,
                database=bd,
                cursorclass=pymysql.cursors.DictCursor
            )
    
    # Para desarrollo local (cuando pruebas en tu PC)
    return pymysql.connect(
        host=os.environ.get('MYSQLHOST', 'localhost'),
        port=int(os.environ.get('MYSQLPORT', 3306)),
        user=os.environ.get('MYSQLUSER', 'root'),
        password=os.environ.get('MYSQLPASSWORD', 'admin123'),
        database=os.environ.get('MYSQLDATABASE', 'railway'),
        cursorclass=pymysql.cursors.DictCursor
    )

def login_requerido(f):
    @wraps(f)
    def decorador(*args, **kwargs):
        if 'usuario' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorador

def admin_requerido(f):
    @wraps(f)
    def decorador(*args, **kwargs):
        if 'usuario' not in session:
            return redirect(url_for('login'))
        if session.get('rol') != 'Administrador':
            return redirect(url_for('acceso_denegado'))
        return f(*args, **kwargs)
    return decorador

@app.route('/')
def index():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('habitaciones'))

@app.route('/acceso_denegado')
def acceso_denegado():
    return render_template('acceso_denegado.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form['usuario']
        clave = request.form['clave']
        conn = conectar()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM USUARIOS WHERE NOMBRE_USUARIO=%s AND CONTRASEÑA_HASH=%s", (usuario, clave))
        user = cursor.fetchone()
        conn.close()
        if user:
            session['usuario'] = usuario
            session['rol'] = user['ROL']
            flash(f'👋 Bienvenido, {usuario}', 'success')
            return redirect(url_for('habitaciones'))
        return render_template('login.html', error='Usuario o contraseña incorrectos')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ─── HABITACIONES ───────────────────────────────────────

@app.route('/habitaciones')
@login_requerido
def habitaciones():
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 10
    offset = (pagina - 1) * por_pagina
    
    busqueda = request.args.get('busqueda', '')
    
    conn = conectar()
    cursor = conn.cursor()
    
    if busqueda:
        cursor.execute("""
            SELECT COUNT(*) as total FROM HABITACIONES 
            WHERE NUMERO LIKE %s OR TIPO LIKE %s
        """, (f'%{busqueda}%', f'%{busqueda}%'))
    else:
        cursor.execute("SELECT COUNT(*) as total FROM HABITACIONES")
    
    total = cursor.fetchone()['total']
    total_paginas = math.ceil(total / por_pagina)
    
    if busqueda:
        cursor.execute("""
            SELECT * FROM HABITACIONES 
            WHERE NUMERO LIKE %s OR TIPO LIKE %s
            ORDER BY NUMERO LIMIT %s OFFSET %s
        """, (f'%{busqueda}%', f'%{busqueda}%', por_pagina, offset))
    else:
        cursor.execute("SELECT * FROM HABITACIONES ORDER BY NUMERO LIMIT %s OFFSET %s", (por_pagina, offset))
    
    habitaciones = cursor.fetchall()
    conn.close()
    
    return render_template('habitaciones.html', 
                         habitaciones=habitaciones, 
                         pagina=pagina, 
                         total_paginas=total_paginas,
                         busqueda=busqueda,
                         total=total)

@app.route('/agregar_habitacion', methods=['POST'])
@login_requerido
def agregar_habitacion():
    numero = request.form['numero']
    tipo = request.form['tipo']
    precio = request.form['precio']
    estado = request.form['estado']
    
    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) as count FROM HABITACIONES WHERE NUMERO=%s", (numero,))
        if cursor.fetchone()['count'] > 0:
            flash('❌ Ya existe una habitación con ese número', 'danger')
            return redirect(url_for('habitaciones'))
        
        cursor.execute(
            "INSERT INTO HABITACIONES (NUMERO, TIPO, PRECIO, ESTADO) VALUES (%s, %s, %s, %s)",
            (numero, tipo, precio, estado)
        )
        conn.commit()
        flash('✅ Habitación agregada correctamente', 'success')
    except Exception as e:
        flash(f'❌ Error al agregar habitación: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('habitaciones'))

@app.route('/editar_habitacion', methods=['POST'])
@login_requerido
def editar_habitacion():
    id_hab = request.form['id']
    numero = request.form['numero']
    tipo = request.form['tipo']
    precio = request.form['precio']
    estado = request.form['estado']
    
    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) as count FROM HABITACIONES WHERE NUMERO=%s AND ID_HABITACION!=%s", (numero, id_hab))
        if cursor.fetchone()['count'] > 0:
            flash('❌ Ya existe otra habitación con ese número', 'danger')
            return redirect(url_for('habitaciones'))
        
        cursor.execute(
            "UPDATE HABITACIONES SET NUMERO=%s, TIPO=%s, PRECIO=%s, ESTADO=%s WHERE ID_HABITACION=%s",
            (numero, tipo, precio, estado, id_hab)
        )
        conn.commit()
        flash('✅ Habitación actualizada correctamente', 'success')
    except Exception as e:
        flash(f'❌ Error al actualizar habitación: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('habitaciones'))

@app.route('/eliminar_habitacion/<int:id>', methods=['POST'])
@admin_requerido
def eliminar_habitacion(id):
    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) as count FROM RESERVAS WHERE ID_HABITACION=%s AND ESTADO='Confirmada'", (id,))
        if cursor.fetchone()['count'] > 0:
            flash('❌ No se puede eliminar la habitación porque tiene reservas activas', 'danger')
            return redirect(url_for('habitaciones'))
        
        cursor.execute("DELETE FROM HABITACIONES WHERE ID_HABITACION=%s", (id,))
        conn.commit()
        flash('🗑️ Habitación eliminada', 'danger')
    except Exception as e:
        flash(f'❌ Error al eliminar habitación: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('habitaciones'))

@app.route('/exportar_habitaciones')
@login_requerido
def exportar_habitaciones():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM HABITACIONES ORDER BY NUMERO")
    datos = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Habitaciones"

    headers = ['N° Habitación', 'Tipo', 'Precio', 'Estado']
    verde = PatternFill("solid", fgColor="1a4a2e")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = verde
        cell.alignment = Alignment(horizontal='center')

    for row, h in enumerate(datos, 2):
        ws.cell(row=row, column=1, value=h['NUMERO'])
        ws.cell(row=row, column=2, value=h['TIPO'])
        ws.cell(row=row, column=3, value=float(h['PRECIO']))
        ws.cell(row=row, column=4, value=h['ESTADO'])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='habitaciones.xlsx', as_attachment=True)

# ─── CLIENTES ───────────────────────────────────────────

@app.route('/clientes')
@login_requerido
def clientes():
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 10
    offset = (pagina - 1) * por_pagina
    
    busqueda = request.args.get('busqueda', '')
    
    conn = conectar()
    cursor = conn.cursor()
    
    if busqueda:
        cursor.execute("""
            SELECT COUNT(*) as total FROM CLIENTES 
            WHERE NOMBRE LIKE %s OR DNI LIKE %s
        """, (f'%{busqueda}%', f'%{busqueda}%'))
    else:
        cursor.execute("SELECT COUNT(*) as total FROM CLIENTES")
    
    total = cursor.fetchone()['total']
    total_paginas = math.ceil(total / por_pagina)
    
    if busqueda:
        cursor.execute("""
            SELECT * FROM CLIENTES 
            WHERE NOMBRE LIKE %s OR DNI LIKE %s
            ORDER BY NOMBRE LIMIT %s OFFSET %s
        """, (f'%{busqueda}%', f'%{busqueda}%', por_pagina, offset))
    else:
        cursor.execute("SELECT * FROM CLIENTES ORDER BY NOMBRE LIMIT %s OFFSET %s", (por_pagina, offset))
    
    clientes = cursor.fetchall()
    conn.close()
    
    return render_template('clientes.html', 
                         clientes=clientes, 
                         pagina=pagina, 
                         total_paginas=total_paginas,
                         busqueda=busqueda,
                         total=total)

@app.route('/registrar_cliente', methods=['POST'])
@login_requerido
def registrar_cliente():
    nombre = request.form['nombre']
    dni = request.form['dni']
    telefono = request.form.get('telefono', '')
    email = request.form.get('email', '')
    
    if not dni.isdigit() or len(dni) != 8:
        flash('❌ DNI debe tener 8 dígitos numéricos', 'danger')
        return redirect(url_for('clientes'))
    
    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) as count FROM CLIENTES WHERE DNI=%s", (dni,))
        if cursor.fetchone()['count'] > 0:
            flash('❌ Ya existe un cliente con ese DNI', 'danger')
            return redirect(url_for('clientes'))
        
        cursor.execute(
            "INSERT INTO CLIENTES (NOMBRE, DNI, TELEFONO, EMAIL) VALUES (%s, %s, %s, %s)",
            (nombre, dni, telefono, email)
        )
        conn.commit()
        flash('✅ Cliente registrado correctamente', 'success')
    except Exception as e:
        flash(f'❌ Error al registrar cliente: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('clientes'))

@app.route('/editar_cliente', methods=['POST'])
@login_requerido
def editar_cliente():
    id_cliente = request.form['id']
    nombre = request.form['nombre']
    dni = request.form['dni']
    telefono = request.form.get('telefono', '')
    email = request.form.get('email', '')
    
    if not dni.isdigit() or len(dni) != 8:
        flash('❌ DNI debe tener 8 dígitos numéricos', 'danger')
        return redirect(url_for('clientes'))
    
    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) as count FROM CLIENTES WHERE DNI=%s AND ID_CLIENTE!=%s", (dni, id_cliente))
        if cursor.fetchone()['count'] > 0:
            flash('❌ Ya existe otro cliente con ese DNI', 'danger')
            return redirect(url_for('clientes'))
        
        cursor.execute(
            "UPDATE CLIENTES SET NOMBRE=%s, DNI=%s, TELEFONO=%s, EMAIL=%s WHERE ID_CLIENTE=%s",
            (nombre, dni, telefono, email, id_cliente)
        )
        conn.commit()
        flash('✅ Cliente actualizado correctamente', 'success')
    except Exception as e:
        flash(f'❌ Error al actualizar cliente: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('clientes'))

@app.route('/eliminar_cliente/<int:id>', methods=['POST'])
@admin_requerido
def eliminar_cliente(id):
    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT COUNT(*) as count FROM RESERVAS WHERE ID_CLIENTE=%s", (id,))
        if cursor.fetchone()['count'] > 0:
            flash('❌ No se puede eliminar el cliente porque tiene reservas asociadas', 'danger')
            return redirect(url_for('clientes'))
        
        cursor.execute("DELETE FROM CLIENTES WHERE ID_CLIENTE=%s", (id,))
        conn.commit()
        flash('🗑️ Cliente eliminado', 'danger')
    except Exception as e:
        flash(f'❌ Error al eliminar cliente: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('clientes'))

@app.route('/exportar_clientes')
@login_requerido
def exportar_clientes():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM CLIENTES")
    datos = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = ['Nombre', 'DNI', 'Teléfono', 'Email']
    verde = PatternFill("solid", fgColor="1a4a2e")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = verde
        cell.alignment = Alignment(horizontal='center')

    for row, c in enumerate(datos, 2):
        ws.cell(row=row, column=1, value=c['NOMBRE'])
        ws.cell(row=row, column=2, value=c['DNI'])
        ws.cell(row=row, column=3, value=c['TELEFONO'])
        ws.cell(row=row, column=4, value=c['EMAIL'])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='clientes.xlsx', as_attachment=True)

# ─── RESERVAS ───────────────────────────────────────────

@app.route('/reservas')
@login_requerido
def reservas():
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 10
    offset = (pagina - 1) * por_pagina
    
    conn = conectar()
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) as total FROM RESERVAS")
    total = cursor.fetchone()['total']
    total_paginas = math.ceil(total / por_pagina)
    
    cursor.execute("""
        SELECT r.ID_RESERVA, c.NOMBRE, h.NUMERO,
               r.FECHA_INGRESO, r.FECHA_SALIDA, r.ESTADO
        FROM RESERVAS r
        JOIN CLIENTES c ON r.ID_CLIENTE = c.ID_CLIENTE
        JOIN HABITACIONES h ON r.ID_HABITACION = h.ID_HABITACION
        ORDER BY r.ID_RESERVA DESC
        LIMIT %s OFFSET %s
    """, (por_pagina, offset))
    reservas = cursor.fetchall()
    
    cursor.execute("SELECT * FROM CLIENTES ORDER BY NOMBRE")
    clientes = cursor.fetchall()
    cursor.execute("SELECT * FROM HABITACIONES WHERE ESTADO = 'Disponible'")
    habitaciones = cursor.fetchall()
    conn.close()
    
    return render_template('reservas.html', 
                         reservas=reservas, 
                         clientes=clientes, 
                         habitaciones=habitaciones,
                         pagina=pagina,
                         total_paginas=total_paginas,
                         total=total)

@app.route('/registrar_reserva', methods=['POST'])
@login_requerido
def registrar_reserva():
    id_cliente = request.form['id_cliente']
    id_habitacion = request.form['id_habitacion']
    fecha_ingreso = request.form['fecha_ingreso']
    fecha_salida = request.form['fecha_salida']
    
    hoy = datetime.now().date()
    ingreso = datetime.strptime(fecha_ingreso, '%Y-%m-%d').date()
    salida = datetime.strptime(fecha_salida, '%Y-%m-%d').date()
    
    if ingreso < hoy:
        flash('❌ La fecha de ingreso no puede ser anterior a hoy', 'danger')
        return redirect(url_for('reservas'))
    
    if salida <= ingreso:
        flash('❌ La fecha de salida debe ser posterior a la fecha de ingreso', 'danger')
        return redirect(url_for('reservas'))
    
    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT COUNT(*) as count FROM RESERVAS 
            WHERE ID_HABITACION=%s AND ESTADO='Confirmada'
            AND ((FECHA_INGRESO <= %s AND FECHA_SALIDA >= %s))
        """, (id_habitacion, fecha_salida, fecha_ingreso))
        
        if cursor.fetchone()['count'] > 0:
            flash('❌ La habitación ya está reservada en esas fechas', 'danger')
            return redirect(url_for('reservas'))
        
        cursor.execute(
            "INSERT INTO RESERVAS (ID_CLIENTE, ID_HABITACION, FECHA_INGRESO, FECHA_SALIDA) VALUES (%s, %s, %s, %s)",
            (id_cliente, id_habitacion, fecha_ingreso, fecha_salida)
        )
        cursor.execute("UPDATE HABITACIONES SET ESTADO='Ocupada' WHERE ID_HABITACION=%s", (id_habitacion,))
        conn.commit()
        flash('✅ Reserva registrada correctamente', 'success')
    except Exception as e:
        flash(f'❌ Error al registrar reserva: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('reservas'))

@app.route('/editar_reserva', methods=['POST'])
@login_requerido
def editar_reserva():
    id_reserva = request.form['id']
    estado = request.form['estado']
    
    conn = conectar()
    cursor = conn.cursor()
    try:
        if estado in ['Cancelada', 'Finalizada']:
            cursor.execute("""
                SELECT ID_HABITACION FROM RESERVAS WHERE ID_RESERVA=%s
            """, (id_reserva,))
            id_habitacion = cursor.fetchone()['ID_HABITACION']
            cursor.execute("UPDATE HABITACIONES SET ESTADO='Disponible' WHERE ID_HABITACION=%s", (id_habitacion,))
        
        cursor.execute(
            "UPDATE RESERVAS SET ESTADO=%s WHERE ID_RESERVA=%s",
            (estado, id_reserva)
        )
        conn.commit()
        flash('✅ Reserva actualizada correctamente', 'success')
    except Exception as e:
        flash(f'❌ Error al actualizar reserva: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('reservas'))

@app.route('/eliminar_reserva/<int:id>', methods=['POST'])
@admin_requerido
def eliminar_reserva(id):
    conn = conectar()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT ID_HABITACION FROM RESERVAS WHERE ID_RESERVA=%s", (id,))
        resultado = cursor.fetchone()
        if resultado:
            id_habitacion = resultado['ID_HABITACION']
            cursor.execute("UPDATE HABITACIONES SET ESTADO='Disponible' WHERE ID_HABITACION=%s", (id_habitacion,))
        
        cursor.execute("DELETE FROM RESERVAS WHERE ID_RESERVA=%s", (id,))
        conn.commit()
        flash('🗑️ Reserva eliminada', 'danger')
    except Exception as e:
        flash(f'❌ Error al eliminar reserva: {e}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('reservas'))

@app.route('/exportar_reservas')
@login_requerido
def exportar_reservas():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT c.NOMBRE, h.NUMERO, r.FECHA_INGRESO, r.FECHA_SALIDA, r.ESTADO
        FROM RESERVAS r
        JOIN CLIENTES c ON r.ID_CLIENTE = c.ID_CLIENTE
        JOIN HABITACIONES h ON r.ID_HABITACION = h.ID_HABITACION
        ORDER BY r.ID_RESERVA DESC
    """)
    datos = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reservas"

    headers = ['Cliente', 'Habitación', 'Fecha Ingreso', 'Fecha Salida', 'Estado']
    verde = PatternFill("solid", fgColor="1a4a2e")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = verde
        cell.alignment = Alignment(horizontal='center')

    for row, r in enumerate(datos, 2):
        ws.cell(row=row, column=1, value=r['NOMBRE'])
        ws.cell(row=row, column=2, value=r['NUMERO'])
        ws.cell(row=row, column=3, value=str(r['FECHA_INGRESO']))
        ws.cell(row=row, column=4, value=str(r['FECHA_SALIDA']))
        ws.cell(row=row, column=5, value=r['ESTADO'])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='reservas.xlsx', as_attachment=True)

# ─── DASHBOARD ──────────────────────────────────────────

@app.route('/dashboard')
@login_requerido
def dashboard():
    conn = conectar()
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) as total FROM HABITACIONES")
    total_habitaciones = cursor.fetchone()['total']
    
    cursor.execute("SELECT COUNT(*) as total FROM HABITACIONES WHERE ESTADO='Disponible'")
    disponibles = cursor.fetchone()['total']
    
    cursor.execute("SELECT COUNT(*) as total FROM HABITACIONES WHERE ESTADO='Ocupada'")
    ocupadas = cursor.fetchone()['total']
    
    cursor.execute("SELECT COUNT(*) as total FROM CLIENTES")
    total_clientes = cursor.fetchone()['total']
    
    cursor.execute("SELECT COUNT(*) as total FROM RESERVAS WHERE ESTADO='Confirmada'")
    reservas_activas = cursor.fetchone()['total']
    
    conn.close()
    
    return render_template('dashboard.html',
                         total_habitaciones=total_habitaciones,
                         disponibles=disponibles,
                         ocupadas=ocupadas,
                         total_clientes=total_clientes,
                         reservas_activas=reservas_activas)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
